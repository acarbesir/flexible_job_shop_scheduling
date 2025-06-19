import numpy as np
import openpyxl as xl
import json
import random

import order
import order_detail
import product
import operation_card
import station

class DataProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.load_data()

    def load_data(self):
        self.orders = self.read_orders()
        self.order_details = self.read_order_details()
        self.products = self.read_products()
        self.operation_cards = self.read_operation_cards()
        self.stations = self.read_stations()

    def read_orders(self):
        wb = xl.load_workbook(self.file_path)
        sheet = wb['orders']
        orders = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            orders.append(order.Order(*row))
        return orders

    def read_order_details(self):
        wb = xl.load_workbook(self.file_path)
        sheet = wb['order_details']
        order_details = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            order_details.append(order_detail.OrderDetail(*row))
        return order_details

    def read_products(self):
        wb = xl.load_workbook(self.file_path)
        sheet = wb['products']
        products = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            products.append(product.Product(*row))
        return products

    def read_operation_cards(self):
        wb = xl.load_workbook(self.file_path)
        sheet = wb['operation_cards']
        operation_cards = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            operation_cards.append(operation_card.OperationCard(*row))
        return operation_cards

    def read_stations(self):
        wb = xl.load_workbook(self.file_path)
        sheet = wb['stations']
        stations = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            station_no, operation, station_name, is_available = row[:4]
            is_available = is_available == "Evet"
            cost_per_minute = random.randint(20, 50)
            stations.append(station.Station(station_no, operation, station_name, is_available, cost_per_minute))
        return stations

    def get_available_machines(self, operation_type, target_cycle_time):
        machines = []
        for station in self.stations:
            if station.operation == operation_type and station.is_available:
                machines.append({"machine_id": int(station.station_no.split('_')[1]),
                                 "target_cycle_time": target_cycle_time})
                
        return machines

    def process(self):
        output_data = {"machines": [], "orders": []}

        for station in self.stations:
            machine_id = int(station.station_no.split('_')[1])
            machine = {
                "machine_id": machine_id,
                "machine_name": station.station_name,
                "machine_availability": station.is_available
            }
            output_data["machines"].append(machine)

        merged_orders = {}

        for order in self.orders:
            if order.to_be_planned == "Evet":
                order_details = [od for od in self.order_details if od.baca_order_no == order.baca_order_no]
                for order_detail in order_details:
                    key = (order_detail.baca_order_no, order_detail.product_no)
                    if key in merged_orders:
                        merged_orders[key]["quantity"] += order_detail.quantity
                    else:
                        deadline = order.baca_due_date.strftime("%d.%m.%Y")
                        merged_orders[key] = {
                            "baca_order_id": order_detail.baca_order_no,
                            "deadline": deadline,
                            "product_code": order_detail.product_no,
                            "quantity": order_detail.quantity,
                            "operations": []
                        }

        for (baca_order_no, product_no), order_data in merged_orders.items():
            product_detail = next((p for p in self.products if p.baca_product_name == product_no), None)
            if not product_detail:
                continue
            product_operations = [op for op in self.operation_cards if op.baca_product_no == product_detail.baca_product_name]
            for operation in product_operations:
                available_machines = self.get_available_machines(operation.operation_type, operation.cycle_time)
                if available_machines:
                    operation_data = {
                        "operation_id": operation.operation_no,
                        "operation_type": operation.operation_type,
                        "setup_time": None,
                        "available_machines": available_machines
                    }
                    order_data["operations"].append(operation_data)

            output_data["orders"].append(order_data)

        return output_data


class JSONExporter:
    @staticmethod
    def export(data, output_path):
        with open(output_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, indent=4, default=JSONExporter.convert_numpy_types)

    @staticmethod
    def convert_numpy_types(obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        return obj


def main(file_path, output):
    processor = DataProcessor(file_path)
    output_data = processor.process()

    JSONExporter.export(output_data, output)
    print(f"Data exported to {output}")

if __name__ == "__main__":
    main()
